import openpyxl
import datetime
from datetime import date, timedelta
from dateutil import parser
otchet = openpyxl.load_workbook('Маспекс Восток 27.03.xlsx')
moscow = openpyxl.load_workbook('Маспекс Восток шаблон27.xlsx')
saint_p = openpyxl.load_workbook('Маспекс-Восток 28.03.2018н.xlsx')

lenta = otchet['Лента']
lenta_m = moscow['Лента']
lenta_s = saint_p['Лента']
result = otchet['Отчёт']

yesterday = date.today() - timedelta(1)
result.cell(row = 2, column = 3, value = yesterday)
#print(yesterday.strftime('%d.%m.%y'))

for i in range(3,35):
    lenta.cell(row=i, column = 15, value = yesterday)

for j in range (35,63):
    lenta.cell(row=j, column = 15, value = lenta_m.cell(row=j-33, column=7).value.date())
lenta.cell(row=65, column=30, value='=COUNTIF(O3:O62'+',"='+ yesterday.strftime('%d.%m.%y')+'"')

for k in range (3,35):
    for m in range(16,29):
        lenta.cell(row=k, column=m, value=lenta_s.cell(row=k,column=m-4).value)

for s in range (2,29):
	for t in range (5,18):
		lenta.cell(row=s, column=t, value= lenta_m.cell(row=s+34, column=t).value)

okay = otchet['Окей']
okay_m = moscow['Окей']
okay_s = saint_p['Окей']



#print(lenta.cell(row=38, column=15).value.date())

#test_parser = parser.parse(lenta.cell(row=38, column=15).value, dayfirst=True)
#print(test_parser)
#lenta.cell(row=7, column =15, value = str(lenta.cell(row=7, column=15).value[1:]))

#for l in range (3,30):
#    print(lenta_m.cell(row=l, column=7).value.strftime("%d.%m.%y"))
otchet.save('test.xlsx')

#print(type(lenta_m.cell(row=3, column=7).value))




#dates_lenta_m=[]
#for l in range (2,30):
    #dates_lenta_m+= datetime.strptime(lenta_m.cell(row=l, column=7).value.strftime("%d.%m.%Y"),'%d.%m.%Y')
    #print(dates_lenta_m)
